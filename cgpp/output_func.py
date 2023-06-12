import logging
from typing import Any, Union

from input_class import ProbInsS21, ProbInsS21T1, ProbInsS21T2, ProbInsS21T3
from meta_class import OutputMetadata
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from output_class import Variables
from santini_21_milp_t2 import solve_santini_21_milp_t2
from santini_21_milp_t3 import solve_santini_21_milp_t3


def make_summary_dict(
    p_ins: ProbInsS21, solution: Variables, output_meta: OutputMetadata
) -> dict[str, Any]:
    return_dict: dict[str, Any] = dict()
    return_dict[output_meta.summary_colhead_list[0]] = p_ins.get_prob_obj_name()
    return_dict[output_meta.summary_colhead_list[1]] = p_ins.n_crops
    return_dict[output_meta.summary_colhead_list[2]] = p_ins.cabinet
    return_dict[output_meta.summary_colhead_list[3]] = p_ins.demand_mult
    return_dict[output_meta.summary_colhead_list[4]] = p_ins.n_days

    return_dict[output_meta.summary_colhead_list[12]] = solution.wall_sec
    return_dict[output_meta.summary_colhead_list[5]] = int(solution.not_solved)
    return_dict[output_meta.summary_colhead_list[6]] = int(solution.is_unbounded)
    return_dict[output_meta.summary_colhead_list[7]] = int(solution.is_infeasible)
    return_dict[output_meta.summary_colhead_list[8]] = int(solution.found_feasible)
    return_dict[output_meta.summary_colhead_list[9]] = int(solution.is_optimal)

    if solution.not_solved or solution.is_unbounded or solution.is_infeasible:
        return_dict[output_meta.summary_colhead_list[13]] = None
        return_dict[output_meta.summary_colhead_list[14]] = None
        return_dict[output_meta.summary_colhead_list[10]] = None
        return_dict[output_meta.summary_colhead_list[11]] = None
        return return_dict

    if solution.found_feasible:
        ub, lb = solution.obj_val, solution.obj_bound
        return_dict[output_meta.summary_colhead_list[13]] = lb
        return_dict[output_meta.summary_colhead_list[14]] = ub
        if solution.is_optimal:
            opt_gap0 = opt_gap1 = 0
        else:
            if lb == 0:
                opt_gap0 = "infty"
            else:
                opt_gap0 = (ub - lb) / lb
            # https://www.gurobi.com/documentation/9.5/refman/mipgap2.html
            opt_gap1 = (ub - lb) / ub
    return_dict[output_meta.summary_colhead_list[10]] = opt_gap0
    return_dict[output_meta.summary_colhead_list[11]] = opt_gap1

    return return_dict


def schedule_by_santini_21_milp_t2(
    p_ins: ProbInsS21T2,
    obj_idx: int,
    solver_name: str,
    timelimit: int,
    output_meta: OutputMetadata,
) -> dict[str, Any]:
    sol = solve_santini_21_milp_t2(p_ins, solver_name, timelimit, obj_idx)
    if sol.found_feasible:
        make_santini_21_sch_xlsx(sol, p_ins, output_meta)
    return make_summary_dict(p_ins, sol, output_meta)


def schedule_by_santini_21_milp_t3(
    p_ins: ProbInsS21T3,
    obj_idx: int,
    solver_name: str,
    timelimit: int,
    output_meta: OutputMetadata,
) -> dict[str, Any]:
    sol = solve_santini_21_milp_t3(p_ins, solver_name, timelimit)
    if sol.found_feasible:
        make_santini_21_sch_xlsx(sol, p_ins, output_meta)
    return make_summary_dict(p_ins, sol, output_meta)


def schedule_by_santini_21_milp(
    p_ins: ProbInsS21,
    model_type_obj_idx_list_dict: dict[str, list[int]],
    solver_name: str,
    timelimit: int,
    output_meta: OutputMetadata,
):
    p_ins.create_t_idx_list()
    for obj_idx in model_type_obj_idx_list_dict[p_ins.model_type]:
        p_ins.set_prob_obj_name(obj_idx)
        logging.info(p_ins.get_prob_obj_name())
        summary_dict = None
        if obj_idx in [1, 2, 4]:
            summary_dict = schedule_by_santini_21_milp_t2(
                p_ins, obj_idx, solver_name, timelimit, output_meta
            )
        elif obj_idx == 3:
            summary_dict = schedule_by_santini_21_milp_t3(
                p_ins, obj_idx, solver_name, timelimit, output_meta
            )
        else:
            continue

        if summary_dict is not None:
            output_meta.append_summary(summary_dict, obj_idx)


def make_santini_21_sch_xlsx(
    solution: Variables, p_ins: ProbInsS21, output_meta: OutputMetadata
):
    # Indices
    C_list = p_ins.crop_id_list  # c\in C in paper
    # S_list = p_ins.shelf_id_list
    D_list = p_ins.t_idx_list[:-1]  # d\in D in paper; set of growth days
    # day 0 is the earliest day seeds can be planted, day 1 is the earliest day of growth counted
    K_list = p_ins.config_id_list  # k\in K in paper
    cfg_id: Union[str, None]
    if type(p_ins) == ProbInsS21T2 or ProbInsS21T3:
        T_list = p_ins.shelf_type_list  # t\in T in paper

    # print(C_list)
    # print(T_list)
    # print(K_list)
    # print(demand_D_list)
    # print(growth_D_list)
    # print(D_prime_list)

    # Parameters
    # c -> the number of growth days
    gamma_dict = p_ins.crop_growth_days
    # c -> g (growth day) -> configuration required
    k_dict = p_ins.crop_growth_day_config_dict
    # c -> t -> the crop can grow on the shelf type
    delta_dict = p_ins.crop_shelf_type_compatible
    # c -> list of compatible t
    T_dict = {c: [t for t in T_list if delta_dict[c][t]] for c in C_list}

    # seed vault index
    sigma = "SeedVault"
    # produce stage index
    tau = "Produce"
    # S_prime_list = [sigma, tau] + S_list
    # S_prime_dict = {c: [sigma, tau] + S_dict[c] for c in C_list}
    for c in C_list:
        delta_dict[c][sigma] = True
        delta_dict[c][tau] = True
    T_prime_dict = {c: [sigma, tau] + T_dict[c] for c in C_list}

    # create empty workbook
    wb = Workbook()

    cursor = {"row": 1, "column": 1}  # each index starts from 1
    # schedule worksheet
    sch_ws: Worksheet = wb.worksheets[0]
    sch_ws.title = output_meta.schedule_sheetname
    # shelf configuration worksheet
    cfg_ws: Worksheet = wb.create_sheet(output_meta.shelf_config_sheetname)
    # germination worksheet
    grm_ws: Worksheet = wb.create_sheet(output_meta.germination_sheetname)
    # harvest worksheet
    hv_ws: Worksheet = wb.create_sheet(output_meta.harvest_sheetname)

    crop_ws_list: list[Worksheet] = [sch_ws, grm_ws, hv_ws]
    shelf_ws_list: list[Worksheet] = [cfg_ws]
    ws_list: list[Worksheet] = crop_ws_list + shelf_ws_list

    cursor["row"] = 1
    time_row = [output_meta.keystone_val] + [0] + p_ins.t_idx_list

    for ws in ws_list:
        ws.cell(**cursor).value = p_ins.get_prob_obj_name()
        cursor["row"] = 2
        ws.freeze_panes = ws["B2"]
        for idx, val in enumerate(time_row):
            col = idx + cursor["column"]
            # define column width
            ws.column_dimensions[get_column_letter(col)].width = output_meta.sch_col_wth
            # write time index row
            ws.cell(column=col, row=cursor["row"]).value = val
        # first column width
        ws.column_dimensions[
            get_column_letter(cursor["column"])
        ].width = output_meta.first_col_wth

    # Shelf configuration worksheet
    cursor["row"] = 2
    # shelf_id -> cfg_id -> list of the number of shelves with the config
    shelf_config_dict: dict[str, dict[str, list[int]]] = {
        shelf_id: {cfg_id: [0] * (p_ins.n_days + 1) for cfg_id in K_list}
        for shelf_id in T_list
    }
    for shelf_id in T_list:
        cursor["row"] += 1
        cfg_ws.cell(**cursor).value = shelf_id

        for t_idx in D_list:
            for cfg_id in K_list:
                cfg_count = int(solution.y[shelf_id][t_idx][cfg_id])
                if cfg_count == 0:
                    continue
                shelf_config_dict[shelf_id][cfg_id][t_idx] = cfg_count

        for cfg_id in K_list:
            count_row_list = shelf_config_dict[shelf_id][cfg_id]
            if sum(count_row_list) == 0:
                continue
            cursor["row"] += 1
            cfg_ws.cell(**cursor).value = cfg_id
            for t_idx, cfg_count in enumerate(count_row_list):
                if cfg_count == 0:
                    continue
                col = (
                    cursor["column"] + 1 + t_idx
                )  # +1 since column t_idx starts from 0
                cfg_ws.cell(column=col, row=cursor["row"]).value = cfg_count

    # Crop information worksheets
    cursor["row"] = 2
    for crop_id in C_list:
        cursor["row"] += 1
        for ws in crop_ws_list:
            ws.cell(**cursor).value = crop_id
        row_key_list: list[str] = list()
        grm_row_list_dict: dict[str, list[int]] = dict()  # row_key -> row_list
        sch_row_list_dict: dict[str, list[int]] = dict()  # row_key -> row_list
        hv_row_list_dict: dict[str, list[int]] = dict()  # row_key -> row_list

        for t_idx in D_list:
            for shelf_id in T_list:
                if shelf_id not in T_prime_dict[crop_id]:
                    continue
                # the crop germinated on the shelf type at t_idx
                # grows from t_idx + 1 to t_idx + gamma_dict[crop_id]
                # harvested at t_idx + gamma_dict[crop_id] + 1
                grm_row_list = [0] * (p_ins.n_days + 1)
                _sch_row_list_dict = dict()
                hv_row_list = [0] * (p_ins.n_days + 1)

                # germination
                grm_unit_count = int(solution.x[crop_id][0][sigma][t_idx][shelf_id])
                cfg_id = k_dict[crop_id][1]
                first_row_key = f"{t_idx}-{cfg_id}-{shelf_id}"
                if grm_unit_count > 0:
                    grm_row_list[t_idx] = grm_unit_count

                # growth
                for g_idx in range(1, gamma_dict[c] + 1):
                    growth_t_idx = t_idx + g_idx
                    if growth_t_idx > p_ins.n_days - 1:
                        break
                    cfg_id = k_dict[crop_id][g_idx]
                    crop_unit_count = sum(
                        int(solution.x[crop_id][g_idx][shelf_id][growth_t_idx][t2])
                        for t2 in T_prime_dict[crop_id]
                    )
                    row_key = f"{t_idx}-{cfg_id}-{shelf_id}"
                    if crop_unit_count > 0:
                        if row_key not in _sch_row_list_dict:
                            _sch_row_list_dict[row_key] = [0] * (p_ins.n_days + 1)
                        _sch_row_list_dict[row_key][growth_t_idx] = crop_unit_count

                # harvest at the due date
                harvest_t_idx = t_idx + gamma_dict[crop_id] + 1
                hv_unit_count = 0
                if harvest_t_idx <= p_ins.n_days - 1:
                    hv_unit_count = int(
                        solution.x[crop_id][gamma_dict[c]][shelf_id][harvest_t_idx - 1][
                            tau
                        ]
                    )
                cfg_id = k_dict[crop_id][gamma_dict[c]]
                last_row_key = f"{t_idx}-{cfg_id}-{shelf_id}"
                if hv_unit_count > 0:
                    hv_row_list[harvest_t_idx] = hv_unit_count

                # do not write rows only with 0
                if (
                    sum(grm_row_list)
                    + sum(sum(value) for value in _sch_row_list_dict.values())
                    + sum(hv_row_list)
                    == 0
                ):
                    continue
                for row_key, row_list in _sch_row_list_dict.items():
                    if row_key not in row_key_list:
                        row_key_list.append(row_key)
                    sch_row_list_dict[row_key] = row_list
                grm_row_list_dict[first_row_key] = grm_row_list
                hv_row_list_dict[last_row_key] = hv_row_list

        for row_key in row_key_list:
            sch_row_list = sch_row_list_dict[row_key]
            if row_key in grm_row_list_dict:
                grm_row_list = grm_row_list_dict[row_key]
            else:
                grm_row_list = [0] * (p_ins.n_days + 1)
            if row_key in hv_row_list_dict:
                hv_row_list = hv_row_list_dict[row_key]
            else:
                hv_row_list = [0] * (p_ins.n_days + 1)

            cursor["row"] += 1
            for ws in crop_ws_list:
                ws.cell(**cursor).value = row_key

            for c_idx, sch_row_val in enumerate(sch_row_list):
                if sch_row_val == 0:
                    continue
                col = cursor["column"] + 1 + c_idx
                sch_ws.cell(column=col, row=cursor["row"]).value = sch_row_val
            for c_idx, grm_row_val in enumerate(grm_row_list):
                if grm_row_val == 0:
                    continue
                col = cursor["column"] + 1 + c_idx
                grm_ws.cell(column=col, row=cursor["row"]).value = grm_row_val
            for c_idx, hv_row_val in enumerate(hv_row_list):
                if hv_row_val == 0:
                    continue
                col = cursor["column"] + 1 + c_idx
                hv_ws.cell(column=col, row=cursor["row"]).value = hv_row_val

    f_loc = output_meta.crop_schedule_f_loc(p_ins.get_prob_obj_name())
    wb.save(filename=f_loc)
