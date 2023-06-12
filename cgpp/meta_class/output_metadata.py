import logging
from dataclasses import dataclass
from pathlib import Path, PurePath

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class OutputMetadata:
    encoding: str
    directory: str
    crop_schedule_fn_prefix: str
    crop_schedule_fn_ext: str

    keystone_val: str
    first_col_wth: int
    sch_col_wth: int

    schedule_sheetname: str
    shelf_config_sheetname: str
    germination_sheetname: str
    harvest_sheetname: str

    summary_fn_prefix: str
    summary_fn_ext: str
    summary_colhead_list: list[str]
    all_record_sheet_name: str
    obj_sheet_name_prefix: str

    def set_base_dir(self, meta_dir: PurePath):
        self._base_dir = meta_dir

    def set_ymdhms(self, ymdhms: str):
        self._ymdhms = ymdhms

    def output_dirname(self) -> str:
        return f"{self.directory}_{self._ymdhms}"

    def crop_schedule_f_loc(self, prob_obj_name: str) -> PurePath:
        fn: str = (
            f"{self.crop_schedule_fn_prefix}_{prob_obj_name}{self.crop_schedule_fn_ext}"
        )

        return_path = PurePath(self._base_dir, self.output_dirname(), fn)
        Path(return_path.parent).mkdir(parents=True, exist_ok=True)
        return return_path

    def summary_file_location(self) -> Path:
        fn: str = f"{self.summary_fn_prefix}_{self._ymdhms}{self.summary_fn_ext}"
        return_path = Path(self._base_dir, self.output_dirname(), fn)
        return_path.parent.mkdir(parents=True, exist_ok=True)
        return return_path

    def init_summary_file(self, obj_idx_list: list[int]):
        sf_path = self.summary_file_location()
        if sf_path.exists():
            logging.info(f"Appending to the existing summary file {sf_path}")
        else:
            # create empty workbook
            wb = Workbook()

            # worksheet with all records
            all_ws: Worksheet = wb.worksheets[0]
            all_ws.title = self.all_record_sheet_name
            all_ws.append(self.summary_colhead_list)
            # worksheet for each objective
            for obj_idx in obj_idx_list:
                ws_name = self.obj_ws_name(obj_idx)
                ws: Worksheet = wb.create_sheet(ws_name)
                ws.append(self.summary_colhead_list)

            f_loc = self.summary_file_location()
            wb.save(filename=f_loc)
            logging.info(f"Created new summary file {sf_path}")

    def obj_ws_name(self, obj_idx: int) -> str:
        return f"{self.obj_sheet_name_prefix}{obj_idx}"

    def append_summary(self, summary_dict: dict[str, float], obj_idx: int):
        val_list = [summary_dict[colhead] for colhead in self.summary_colhead_list]

        sf_path = self.summary_file_location()
        wb = load_workbook(sf_path)

        ws = wb[self.all_record_sheet_name]
        ws.append(val_list)
        ws = wb[self.obj_ws_name(obj_idx)]
        ws.append(val_list)

        wb.save(filename=sf_path)
